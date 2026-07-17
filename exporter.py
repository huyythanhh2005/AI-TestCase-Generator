"""
Giai đoạn 7 của pipeline: Export (Excel / Jira / TestRail).

- ExcelExporter: đường xuất chính, 3 sheet (Tóm tắt / Yêu cầu / Test Cases).
- JiraExporter: đẩy mỗi test case thành 1 issue qua Jira Cloud REST API
  v3 (cần JIRA_URL / JIRA_EMAIL / JIRA_API_TOKEN / JIRA_PROJECT_KEY
  trong .env).
- TestRailExporter: đẩy mỗi test case thành 1 case (nhóm theo section =
  module/feature) qua TestRail REST API v2 (cần TESTRAIL_URL /
  TESTRAIL_EMAIL / TESTRAIL_API_KEY / TESTRAIL_PROJECT_ID trong .env).
- export_json(): xuất payload JSON trung gian, dùng khi muốn tự viết
  script import riêng thay vì gọi thẳng 2 exporter trên.
"""

from datetime import datetime

import requests
from requests.auth import HTTPBasicAuth
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import config
from utils import save_json, info, warning


class ExcelExporter:

    def __init__(self):
        self.output_dir = config.OUTPUT_DIR

    def export(self, requirements, testcases, use_cases=None):
        use_cases = use_cases or []

        wb = Workbook()
        wb.remove(wb.active)

        self._write_summary(wb.create_sheet("Tóm tắt"), requirements, testcases, use_cases)
        self._write_use_cases(wb.create_sheet("Use Case"), use_cases)
        self._write_requirements(wb.create_sheet("Yêu cầu"), requirements)
        self._write_testcases(wb.create_sheet("Test Cases"), testcases)

        timestamp = datetime.now().strftime(config.DATE_FORMAT)
        filename = f"{config.EXCEL_NAME}_{timestamp}.xlsx"
        filepath = self.output_dir / filename

        wb.save(filepath)
        info(f"Đã xuất Excel: {filepath}")
        return str(filepath)

    # ------------------------------------------------------------
    def _header_style(self, cell):
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _auto_width(self, sheet):
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value is not None:
                    val = (
                        cell.value.strftime("%Y-%m-%d %H:%M:%S")
                        if isinstance(cell.value, datetime)
                        else str(cell.value)
                    )
                    max_length = max(max_length, len(val))
            sheet.column_dimensions[column_letter].width = min(max(max_length + 3, 12), 65)

    def _write_summary(self, sheet, requirements, testcases, use_cases=None):
        use_cases = use_cases or []

        sheet["A1"] = "BÁO CÁO KIỂM THỬ TỰ ĐỘNG"
        sheet["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        sheet["A1"].fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        sheet.merge_cells("A1:C1")

        row = 3
        for col_idx, header in enumerate(["Hạng mục", "Số lượng", "Ghi chú"], 1):
            self._header_style(sheet.cell(row=row, column=col_idx, value=header))

        metrics = [
            ("Tổng số yêu cầu hệ thống", len(requirements), "Feature Extraction + LLM"),
            ("Tổng số Use Case", len(use_cases), "Use Case Synthesis (gom yêu cầu liên quan)"),
            ("Tổng số kịch bản (Test Cases)", len(testcases), "Sinh bởi AI"),
            ("Thời gian khởi tạo tệp", datetime.now(), "Real-time"),
        ]

        for metric in metrics:
            row += 1
            sheet.cell(row=row, column=1, value=metric[0]).font = Font(bold=True)
            sheet.cell(row=row, column=2, value=metric[1])
            sheet.cell(row=row, column=3, value=metric[2])

        self._auto_width(sheet)

    def _write_use_cases(self, sheet, use_cases):
        headers = [
            "Mã UC", "Tên Use Case", "Actor", "Module", "Mô tả",
            "Điều kiện tiên quyết", "Luồng chính", "Luồng thay thế",
            "Kết quả sau khi hoàn tất", "Tính năng liên quan (feature thật)",
        ]
        for col, header in enumerate(headers, 1):
            self._header_style(sheet.cell(row=1, column=col, value=header))

        def _join(values):
            values = values or []
            if not isinstance(values, list):
                return str(values)
            return "\n".join(f"- {v}" for v in values)

        def _join_flow(values):
            values = values or []
            if not isinstance(values, list):
                return str(values)
            return "\n".join(f"{i + 1}. {v}" for i, v in enumerate(values))

        for row, uc in enumerate(use_cases, 2):
            sheet.cell(row=row, column=1, value=uc.get("uc_id", ""))
            sheet.cell(row=row, column=2, value=uc.get("title", ""))
            sheet.cell(row=row, column=3, value=uc.get("actor", ""))
            sheet.cell(row=row, column=4, value=uc.get("module", ""))
            sheet.cell(row=row, column=5, value=uc.get("description", ""))
            sheet.cell(row=row, column=6, value=_join(uc.get("preconditions")))
            sheet.cell(row=row, column=7, value=_join_flow(uc.get("main_flow")))
            sheet.cell(row=row, column=8, value=_join_flow(uc.get("alternative_flows")))
            sheet.cell(row=row, column=9, value=_join(uc.get("postconditions")))
            sheet.cell(row=row, column=10, value=", ".join(uc.get("related_features") or []))

            for col in range(1, 11):
                sheet.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

        self._auto_width(sheet)

    def _write_requirements(self, sheet, requirements):
        headers = [
            "Tính năng", "Phân vùng Module", "Mô tả", "Quy tắc nghiệp vụ",
            "Quy tắc xác thực", "Bằng chứng từ website (elements thật)",
        ]
        for col, header in enumerate(headers, 1):
            self._header_style(sheet.cell(row=1, column=col, value=header))

        for row, req in enumerate(requirements, 2):
            sheet.cell(row=row, column=1, value=req.get("feature", ""))
            sheet.cell(row=row, column=2, value=req.get("module", ""))
            sheet.cell(row=row, column=3, value=req.get("description", ""))

            b_rules = req.get("business_rules") or []
            v_rules = req.get("validation_rules") or []

            sheet.cell(row=row, column=4, value="\n".join(b_rules) if isinstance(b_rules, list) else str(b_rules))
            sheet.cell(row=row, column=5, value="\n".join(v_rules) if isinstance(v_rules, list) else str(v_rules))
            sheet.cell(row=row, column=6, value=self._summarize_elements(req.get("elements")))

            for col in range(1, 7):
                sheet.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

        self._auto_width(sheet)

    @staticmethod
    def _summarize_elements(elements):
        """Tóm tắt ngắn gọn các element thật (name/id/href/text...) đi
        kèm requirement, để người review Excel có thể tự đối chiếu ngược
        lại với website thật thay vì chỉ tin vào mô tả do AI viết."""
        if not elements or not isinstance(elements, list):
            return ""

        lines = []
        for el in elements[:10]:
            if not isinstance(el, dict):
                continue
            parts = [
                f"{k}={v}" for k, v in el.items()
                if v not in (None, "", [], False) and k != "index"
            ]
            if parts:
                lines.append(", ".join(parts))

        extra = len(elements) - len(lines)
        text = "\n".join(lines)
        if extra > 0:
            text += f"\n... (+{extra} khác)"
        return text

    def _write_testcases(self, sheet, testcases):
        headers = [
            "Mã TC", "Tính năng", "Kịch bản", "Phân loại", "Mức ưu tiên",
            "Các bước", "Kết quả mong đợi", "Element tham chiếu (locator thật)",
            "Use Case liên quan",
        ]
        for col, header in enumerate(headers, 1):
            self._header_style(sheet.cell(row=1, column=col, value=header))

        for row, tc in enumerate(testcases, 2):
            sheet.cell(row=row, column=1, value=tc.get("tc_id", f"TC_{row - 1:05d}"))
            sheet.cell(row=row, column=2, value=tc.get("feature", ""))
            sheet.cell(row=row, column=3, value=tc.get("scenario", ""))
            sheet.cell(row=row, column=4, value=tc.get("type", "Functional"))
            sheet.cell(row=row, column=5, value=tc.get("priority", "Medium"))

            steps = tc.get("steps") or []
            steps_text = (
                "\n".join(f"{i + 1}. {s}" for i, s in enumerate(steps))
                if isinstance(steps, list)
                else str(steps)
            )
            sheet.cell(row=row, column=6, value=steps_text)
            
            # FIX: Convert expected_result to string (handle list case)
            expected_result = tc.get("expected_result", "")
            if isinstance(expected_result, list):
                expected_result_text = "\n".join(f"- {er}" for er in expected_result)
            else:
                expected_result_text = str(expected_result) if expected_result else ""
            sheet.cell(row=row, column=7, value=expected_result_text)
            
            sheet.cell(row=row, column=8, value=tc.get("element_ref", ""))
            sheet.cell(row=row, column=9, value=tc.get("use_case_ref", ""))

            for col in range(1, 10):
                sheet.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

        self._auto_width(sheet)


class JiraExporter:
    """Đẩy mỗi test case thành 1 issue trên Jira Cloud (REST API v3).

    Yêu cầu cấu hình trong .env: JIRA_URL, JIRA_EMAIL, JIRA_API_TOKEN
    (Atlassian API token, KHÔNG phải mật khẩu), JIRA_PROJECT_KEY.
    JIRA_ISSUE_TYPE mặc định là "Test" (một số project Jira dùng loại
    issue riêng cho test case, vd qua add-on Xray/Zephyr - đổi tên loại
    issue nếu cần).
    """

    def __init__(self):
        self.base_url = (config.JIRA_URL or "").rstrip("/")
        self.auth = HTTPBasicAuth(config.JIRA_EMAIL, config.JIRA_API_TOKEN)
        self.project_key = config.JIRA_PROJECT_KEY
        self.issue_type = config.JIRA_ISSUE_TYPE

    def export(self, requirements, testcases):
        if not (self.base_url and config.JIRA_EMAIL and config.JIRA_API_TOKEN and self.project_key):
            raise RuntimeError(
                "Thiếu cấu hình Jira. Cần khai báo đủ JIRA_URL, JIRA_EMAIL, "
                "JIRA_API_TOKEN, JIRA_PROJECT_KEY trong file .env."
            )

        created, failed = [], []

        for tc in testcases:
            payload = self._build_issue_payload(tc)
            try:
                resp = requests.post(
                    f"{self.base_url}/rest/api/3/issue",
                    json=payload,
                    auth=self.auth,
                    headers={"Content-Type": "application/json"},
                    timeout=30,
                )
                if resp.status_code in (200, 201):
                    created.append(resp.json().get("key"))
                else:
                    failed.append({"tc_id": tc.get("tc_id"), "error": resp.text[:300]})
            except Exception as e:
                failed.append({"tc_id": tc.get("tc_id"), "error": str(e)})

        info(f"Jira: đã tạo {len(created)}/{len(testcases)} issue ({len(failed)} lỗi).")
        if failed:
            warning(f"Jira: {len(failed)} test case đẩy thất bại, xem log để biết chi tiết.")

        return {"created_keys": created, "failed": failed}

    def _build_issue_payload(self, tc):
        steps = tc.get("steps") or []
        steps_text = "\n".join(f"{i + 1}. {s}" for i, s in enumerate(steps))

        expected_result = tc.get("expected_result", "")
        if isinstance(expected_result, list):
            expected_result_text = "\n".join(f"- {er}" for er in expected_result)
        else:
            expected_result_text = str(expected_result) if expected_result else ""

        description_text = (
            f"Module: {tc.get('module', '')}\n"
            f"Precondition: {tc.get('precondition', '')}\n"
            f"Test data: {tc.get('test_data', '')}\n"
            f"Element tham chiếu (locator thật): {tc.get('element_ref', '')}\n\n"
            f"Steps:\n{steps_text}\n\n"
            f"Expected result:\n{expected_result_text}"
        )

        return {
            "fields": {
                "project": {"key": self.project_key},
                "summary": f"[{tc.get('type', 'Functional')}] {tc.get('scenario') or tc.get('tc_id', '')}",
                "issuetype": {"name": self.issue_type},
                "description": {
                    "type": "doc",
                    "version": 1,
                    "content": [
                        {
                            "type": "paragraph",
                            "content": [{"type": "text", "text": description_text}],
                        }
                    ],
                },
                "labels": [str(tc.get("type", "Functional")).replace(" ", "_")],
            }
        }


class TestRailExporter:
    """Đẩy test case lên TestRail (REST API v2), gom theo section =
    module/feature (tự tạo section nếu chưa tồn tại).

    Yêu cầu cấu hình trong .env: TESTRAIL_URL, TESTRAIL_EMAIL,
    TESTRAIL_API_KEY, TESTRAIL_PROJECT_ID. TESTRAIL_SUITE_ID chỉ cần
    khi project TestRail bật chế độ nhiều suite.
    """

    def __init__(self):
        self.base_url = (config.TESTRAIL_URL or "").rstrip("/")
        self.auth = HTTPBasicAuth(config.TESTRAIL_EMAIL, config.TESTRAIL_API_KEY)
        self.project_id = config.TESTRAIL_PROJECT_ID
        self.suite_id = config.TESTRAIL_SUITE_ID

    def export(self, requirements, testcases):
        if not (self.base_url and config.TESTRAIL_EMAIL and config.TESTRAIL_API_KEY and self.project_id):
            raise RuntimeError(
                "Thiếu cấu hình TestRail. Cần khai báo đủ TESTRAIL_URL, "
                "TESTRAIL_EMAIL, TESTRAIL_API_KEY, TESTRAIL_PROJECT_ID "
                "trong file .env."
            )

        section_cache = {}
        created, failed = [], []

        for tc in testcases:
            module_name = tc.get("module") or tc.get("feature") or "General"
            try:
                section_id = self._get_or_create_section(module_name, section_cache)
                case_id = self._create_case(section_id, tc)
                created.append(case_id)
            except Exception as e:
                failed.append({"tc_id": tc.get("tc_id"), "error": str(e)})

        info(f"TestRail: đã tạo {len(created)}/{len(testcases)} case ({len(failed)} lỗi).")
        if failed:
            warning(f"TestRail: {len(failed)} test case đẩy thất bại, xem log để biết chi tiết.")

        return {"created_case_ids": created, "failed": failed}

    def _get_or_create_section(self, name, cache):
        if name in cache:
            return cache[name]

        suite_qs = f"&suite_id={self.suite_id}" if self.suite_id else ""
        resp = requests.get(
            f"{self.base_url}/index.php?/api/v2/get_sections/{self.project_id}{suite_qs}",
            auth=self.auth,
            timeout=30,
        )
        resp.raise_for_status()
        payload = resp.json()
        sections = payload.get("sections", []) if isinstance(payload, dict) else payload

        for section in sections:
            if section.get("name") == name:
                cache[name] = section["id"]
                return section["id"]

        body = {"name": name}
        if self.suite_id:
            body["suite_id"] = self.suite_id

        resp = requests.post(
            f"{self.base_url}/index.php?/api/v2/add_section/{self.project_id}",
            json=body,
            auth=self.auth,
            timeout=30,
        )
        resp.raise_for_status()
        section_id = resp.json()["id"]
        cache[name] = section_id
        return section_id

    def _create_case(self, section_id, tc):
        steps = tc.get("steps") or []
        
        expected_result = tc.get("expected_result", "")
        if isinstance(expected_result, list):
            expected_result_text = "\n".join(f"- {er}" for er in expected_result)
        else:
            expected_result_text = str(expected_result) if expected_result else ""
        
        body = {
            "title": tc.get("scenario") or tc.get("tc_id") or "Untitled test case",
            "type_id": 1,
            "priority_id": self._map_priority(tc.get("priority")),
            "custom_preconds": (
                f"{tc.get('precondition', '')}\n"
                f"Element tham chiếu: {tc.get('element_ref', '')}"
            ).strip(),
            "custom_steps": "\n".join(f"{i + 1}. {s}" for i, s in enumerate(steps)),
            "custom_expected": expected_result_text,
            "refs": tc.get("tc_id", ""),
        }
        resp = requests.post(
            f"{self.base_url}/index.php?/api/v2/add_case/{section_id}",
            json=body,
            auth=self.auth,
            timeout=30,
        )
        resp.raise_for_status()
        return resp.json()["id"]

    @staticmethod
    def _map_priority(priority):
        return {"Low": 1, "Medium": 2, "High": 3, "Critical": 4}.get(priority, 2)


def export_json(requirements, testcases, use_cases=None, filename=None):
    """Xuất payload JSON trung gian, sẵn sàng để một script riêng đẩy
    lên Jira/TestRail qua REST API (điểm mở rộng, chưa tự động đẩy)."""

    timestamp = datetime.now().strftime(config.DATE_FORMAT)
    filename = filename or f"testcase_export_{timestamp}.json"
    filepath = config.OUTPUT_DIR / filename

    payload = {
        "generated_at": datetime.now().isoformat(),
        "use_cases": use_cases or [],
        "requirements": requirements,
        "testcases": testcases,
    }

    save_json(payload, filepath)
    info(f"Đã xuất JSON (Jira/TestRail staging): {filepath}")
    return str(filepath)
